"""Importa MP-INS-MO da planilha 'Samuel ORÇAMENTO PADRÃO' para a tabela material.

Padrão de SKU importado: CFxxxSFxxxUxxx (codifica origem da planilha).
Frontend pinta linhas com esse padrão para indicar visualmente a fonte.

Uso:
  python -m scripts.import_planilha_samuel              # dry-run
  python -m scripts.import_planilha_samuel --apply      # aplica ao DB
  python -m scripts.import_planilha_samuel --planilha caminho.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

from app.lib.supabase import get_admin_client

DEFAULT_PLANILHA = Path(
    r"C:\Users\luc98\Documents\Hoje_atualizar\Empresas\0-STEEL\0-METALFORT\0-Samuel ORÇAMENTO PADRÃO.xlsx"
)
SHEET_NAME = "MP-INS-MO"

# Famílias da planilha → categoria do ERP.
FAMILIA_PARA_CATEGORIA = {
    "CF001": "estrutura",
    "CF002": "estrutura",
    "CF003": "estrutura",
    "CF004": "estrutura",
    "CF005": "estrutura",      # parafusos
    "CF006": "servico",        # mão de obra
    "CF008": "fechamento",     # placas cimentícias
    "CF009": "estrutura",      # elementos estruturais
    "CF010": "fechamento",     # desempenho estrutural (banda, manta, resina)
    "CF011": "fechamento",     # vedações (telhas, calhas, membrana)
}

# Perfis sempre vão como metro (mesmo que a planilha indique KG por metro).
PERFIL_FAMILIES = {"CF001", "CF002", "CF003", "CF004"}

# Unidades válidas no ERP (espelha INTEGER_UNITS + numéricas).
UNIDADES_ERP = {"kg", "m", "m2", "pc", "cx", "und", "h", "bd", "rl", "sc", "ml", "ct"}

UNIDADE_PLANILHA_PARA_ERP = {
    "kg": "kg", "un": "und", "m2": "m2", "m": "m",
    "rl": "rl", "bd": "bd", "sc": "sc",
}


def map_unidade(familia: str, unidade_raw) -> str:
    if familia in PERFIL_FAMILIES:
        return "m"
    u = (str(unidade_raw or "")).strip().lower()
    return UNIDADE_PLANILHA_PARA_ERP.get(u, u)


def parse_custo(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_planilha(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' nao encontrada. Sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for raw_row in rows[1:]:  # skip header
        row = list(raw_row) + [None] * (10 - len(raw_row))
        sku, _u_code, _u_name, fam, _fam_desc, _seq, desc_seq, unidade_raw, qtde, custo = row[:10]
        if not sku or not str(sku).strip():
            continue
        if not desc_seq or not str(desc_seq).strip():
            continue
        custo_num = parse_custo(custo)
        if custo_num is None:
            continue
        familia = (str(fam).strip() if fam else "")
        items.append({
            "sku": str(sku).strip(),
            "nome": str(desc_seq).strip(),
            "familia": familia,
            "unidade_planilha_raw": unidade_raw,
            "qtde_un_planilha": qtde,
            "unidade": map_unidade(familia, unidade_raw),
            "preco_unitario": custo_num,
        })
    return items


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--apply", action="store_true", help="Aplica ao DB. Sem flag = dry-run.")
    ap.add_argument("--planilha", default=str(DEFAULT_PLANILHA))
    args = ap.parse_args()

    path = Path(args.planilha)
    if not path.exists():
        raise SystemExit(f"Planilha nao encontrada: {path}")

    items = load_planilha(path)
    if not items:
        print("Nenhum item valido encontrado.")
        sys.exit(1)

    sb = get_admin_client()
    skus = list({i["sku"] for i in items})
    existing_rows = (
        sb.table("material").select("sku, ativo").in_("sku", skus).execute().data or []
    )
    existing_skus = {r["sku"] for r in existing_rows}

    creates: list[dict] = []
    updates: list[dict] = []
    skips: list[tuple[str, str]] = []

    for it in items:
        cat = FAMILIA_PARA_CATEGORIA.get(it["familia"])
        if cat is None:
            skips.append((it["sku"], f"familia '{it['familia']}' sem mapeamento"))
            continue
        unidade = it["unidade"]
        if unidade not in UNIDADES_ERP:
            skips.append((it["sku"], f"unidade '{unidade}' invalida (planilha: {it['unidade_planilha_raw']!r})"))
            continue
        payload = {
            "sku": it["sku"],
            "nome": it["nome"],
            "categoria": cat,
            "unidade": unidade,
            "preco_unitario": it["preco_unitario"],
        }
        (updates if it["sku"] in existing_skus else creates).append(payload)

    cat_count: dict[str, int] = {}
    un_count: dict[str, int] = {}
    for p in creates + updates:
        cat_count[p["categoria"]] = cat_count.get(p["categoria"], 0) + 1
        un_count[p["unidade"]] = un_count.get(p["unidade"], 0) + 1

    print("\n=== Importacao MP-INS-MO ===")
    print(f"Total validos na planilha: {len(items)}")
    print(f"Vai criar: {len(creates)}")
    print(f"Vai atualizar: {len(updates)}")
    print(f"Pular: {len(skips)}")
    if cat_count:
        print("\nCategorias destino:")
        for k, v in sorted(cat_count.items()):
            print(f"  {k:14} {v}")
    if un_count:
        print("\nUnidades destino:")
        for k, v in sorted(un_count.items()):
            print(f"  {k:6} {v}")
    if skips:
        print("\nPulados:")
        for sku, motivo in skips:
            print(f"  {sku}: {motivo}")

    print("\nPreview (todos os criar/atualizar):")
    for p in creates + updates:
        marca = "+" if p in creates else "~"
        print(f"  {marca} {p['sku']:20} {p['nome'][:55]:55} | {p['categoria']:11} | {p['unidade']:4} | R$ {p['preco_unitario']:>10.4f}")

    if not args.apply:
        print("\n[DRY-RUN] Sem alteracoes. Use --apply para aplicar.")
        return

    n_created = 0
    n_updated = 0
    for p in creates:
        sb.table("material").insert({**p, "ativo": True, "estoque_minimo": 0}).execute()
        n_created += 1
    for p in updates:
        sb.table("material").update({**p, "ativo": True}).eq("sku", p["sku"]).execute()
        n_updated += 1
    print(f"\n[APPLY] Criados: {n_created}. Atualizados: {n_updated}.")


if __name__ == "__main__":
    main()
