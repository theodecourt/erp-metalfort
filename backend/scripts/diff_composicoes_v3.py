"""Mapeia composicoes da planilha v3 vs combos atuais do ERP.

Aba COMPOSIÇÕES da v3:
  - linhas "header": tem COMPxxxxx no col 0
  - linhas "filho": col 0 vazio, mas tem material_id no col 6 (Codigo MP-INS-MO)

Saida:
  - lista de composicoes (codigo, descricao, unidade, custo_total, n_itens)
  - combos atuais no ERP (slug, nome, categoria, n_materiais)
  - tentativa de match
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from app.lib.supabase import get_admin_client

PATH = Path(
    r"C:\Users\luc98\Documents\Hoje_atualizar\Empresas\0-STEEL\0-METALFORT"
    r"\erp-metalfort\planilhas tito\0-Samuel ORÇAMENTO PADRÃO v3.xlsx"
)


def main() -> None:
    if not PATH.exists():
        raise SystemExit(f"Planilha nao encontrada: {PATH}")
    wb = openpyxl.load_workbook(PATH, data_only=True)
    ws = wb["COMPOSIÇÕES"]

    # Cabeçalho da aba
    print("=== Cabeçalho da aba COMPOSIÇÕES ===")
    for i, c in enumerate(ws[1]):
        print(f"  col {i}: {c.value!r}")
    print()

    composicoes: list[dict] = []
    current: dict | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and current is None:
            continue
        col0 = str(row[0]).strip() if row[0] else ""
        if col0.startswith("COMP"):
            # Linha header: nova composição
            if current is not None:
                composicoes.append(current)
            current = {
                "codigo": col0,
                "descricao": (str(row[1]).strip() if row[1] else ""),
                "unidade": (str(row[4]).strip().lower() if row[4] else ""),
                "qtd": row[5],
                "custo_total": row[10] if len(row) > 10 else None,
                "itens": [],
            }
        else:
            # Linha filho — pertence ao current
            if current is None:
                continue
            mat_codigo = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            if not mat_codigo:
                continue
            current["itens"].append({
                "material_codigo": mat_codigo,
                "descricao": str(row[7]).strip() if len(row) > 7 and row[7] else "",
                "quantidade": row[8] if len(row) > 8 else None,
                "custo_item": row[9] if len(row) > 9 else None,
            })

    if current is not None:
        composicoes.append(current)

    print(f"=== Composições da v3 ({len(composicoes)}) ===")
    for c in composicoes:
        custo = c["custo_total"]
        custo_str = f"R$ {float(custo):.2f}" if isinstance(custo, (int, float)) else "—"
        print(f"  {c['codigo']:11} | {c['descricao'][:55]:55} | {c['unidade']:4} | {custo_str:>14} | {len(c['itens'])} itens")

    # ERP atual
    sb = get_admin_client()
    combos = (
        sb.table("pacote_combo")
        .select("id, slug, categoria, nome, ativo")
        .order("categoria")
        .order("ordem")
        .execute()
        .data
        or []
    )

    print(f"\n=== Combos atuais no ERP ({len(combos)}) ===")
    by_cat: dict[str, list[dict]] = {}
    for c in combos:
        by_cat.setdefault(c["categoria"], []).append(c)
    for cat in sorted(by_cat):
        print(f"\n  {cat}:")
        for c in by_cat[cat]:
            ativo = "ativo" if c["ativo"] else "INATIVO"
            print(f"    - {c['slug']:30} | {c['nome']:35} | {ativo}")

    # Conta materiais por combo
    mats = sb.table("pacote_combo_material").select("pacote_combo_id, material_id").execute().data or []
    mats_by_combo: dict[str, int] = {}
    for m in mats:
        mats_by_combo[m["pacote_combo_id"]] = mats_by_combo.get(m["pacote_combo_id"], 0) + 1
    total_combo_mats = sum(mats_by_combo.values())
    print(f"\n  Total de combo-material no ERP: {total_combo_mats} ({len(combos)} combos)")

    # Heuristica de match: por descricao
    print(f"\n=== Match heuristico (por substring de descricao) ===")
    descricoes_combos = [(c, (c["nome"] or "").lower()) for c in combos]
    for comp in composicoes:
        desc = comp["descricao"].lower()
        # Tentativa: ver se alguma palavra-chave da composicao bate com algum combo
        match = None
        for c, nome in descricoes_combos:
            # Match grosseiro: se >50% das palavras de pelo menos 4 chars batem
            words_comp = {w for w in desc.split() if len(w) >= 4}
            words_combo = {w for w in nome.split() if len(w) >= 4}
            if not words_comp or not words_combo:
                continue
            common = words_comp & words_combo
            if len(common) >= 2:
                match = c
                break
        if match:
            print(f"  {comp['codigo']:11} {comp['descricao'][:50]:50} ~ {match['slug']:30} ({match['nome'][:25]})")
        else:
            print(f"  {comp['codigo']:11} {comp['descricao'][:50]:50} -> sem match")


if __name__ == "__main__":
    main()
